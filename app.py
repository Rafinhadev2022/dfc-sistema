from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, make_response, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Category, Transaction, Contract, Projection, PasswordResetToken, CostCenter, BillReminder
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from fpdf import FPDF
from sqlalchemy import func, extract
import json, os, io, csv, mimetypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dfc-sistema-chave-secreta-2024')

# Banco de dados: PostgreSQL em produção, SQLite local
DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

if DATABASE_URL:
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    # SQLite: usa /tmp em produção (Render) ou pasta local
    db_path = os.path.join('/tmp', 'dfc.db') if os.environ.get('RENDER') else os.path.join(app.instance_path, 'dfc.db')
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 15 * 1024 * 1024  # 15 MB máximo

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Faça login para acessar o sistema.'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('Acesso restrito ao administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def run_migrations():
    """Garante que todas as colunas necessárias existem. Roda SEMPRE, antes do seed."""
    from sqlalchemy import inspect as sa_inspect, text
    inspector = sa_inspect(db.engine)
    is_pg = 'postgresql' in str(app.config['SQLALCHEMY_DATABASE_URI'])
    blob_type = 'BYTEA' if is_pg else 'BLOB'

    def add_col(table, col, col_type):
        try:
            # Verifica se tabela existe antes de inspecionar colunas
            if table not in inspector.get_table_names():
                return
            cols = {c['name'] for c in inspector.get_columns(table)}
            if col not in cols:
                db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}'))
                db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f'[MIGRAÇÃO] add_col {table}.{col}: {e}')

    # users
    add_col('users', 'active', 'BOOLEAN DEFAULT 1')

    # transactions
    add_col('transactions', 'attachment_data',     blob_type)
    add_col('transactions', 'attachment_original', 'VARCHAR(255)')
    add_col('transactions', 'attachment_mimetype', 'VARCHAR(100)')
    add_col('transactions', 'cost_center_id',      'INTEGER')

    # Remove coluna legada filesystem se existir (PostgreSQL)
    try:
        if 'transactions' in inspector.get_table_names():
            existing_t = {c['name'] for c in inspector.get_columns('transactions')}
            if 'attachment_filename' in existing_t and is_pg:
                db.session.execute(text('ALTER TABLE transactions DROP COLUMN attachment_filename'))
                db.session.commit()
    except Exception:
        db.session.rollback()


def init_database():
    """Cria tabelas e popula dados padrão."""
    db.create_all()
    run_migrations()  # ← migrações logo após create_all, antes de qualquer query

    ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'contatonuees@gmail.com')
    ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'nunes2025#')

    admin = User.query.filter_by(role='admin').first()
    if admin:
        admin.email = ADMIN_EMAIL
        admin.password_hash = generate_password_hash(ADMIN_PASSWORD, method='pbkdf2:sha256')
    else:
        db.session.add(User(
            name='Administrador',
            email=ADMIN_EMAIL,
            password_hash=generate_password_hash(ADMIN_PASSWORD, method='pbkdf2:sha256'),
            role='admin'
        ))
    categorias_entrada = ['Recebimento de Medição','Antecipação de Contrato',
        'Adiantamento','Reembolso de Despesas','Outras Receitas']
    categorias_saida = ['Folha de Pagamento','INSS / FGTS','ISS','IRPJ / CSLL',
        'PIS / COFINS','CBUQ / Massa Asfáltica','Emulsão Asfáltica','Combustível',
        'Manutenção de Equipamentos','Aluguel de Equipamentos','Subempreiteiros',
        'Material de Construção','Sinalização / EPI','Despesas Administrativas',
        'Aluguel / Sede','Contador / Assessoria','Outras Despesas']
    for name in categorias_entrada:
        if not Category.query.filter_by(name=name, type='entrada').first():
            db.session.add(Category(name=name, type='entrada', active=True))
    for name in categorias_saida:
        if not Category.query.filter_by(name=name, type='saida').first():
            db.session.add(Category(name=name, type='saida', active=True))
    db.session.commit()


with app.app_context():
    db.create_all()        # cria tabelas novas sem depender do seed
    run_migrations()       # adiciona colunas faltantes
    try:
        init_database()    # popula dados (admin, categorias)
    except Exception as e:
        print(f'[AVISO] Erro ao inicializar dados: {e}')

# ─── HEALTH CHECK ────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    import traceback
    try:
        db.session.execute(db.text('SELECT 1'))
        user_count = User.query.count()
        return jsonify({
            'status': 'ok',
            'db': str(app.config['SQLALCHEMY_DATABASE_URI']),
            'users': user_count,
            'render_env': os.environ.get('RENDER','not set')
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e), 'trace': traceback.format_exc()}), 500

# ─── AUTH ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('E-mail ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    hoje = date.today()
    inicio_mes = hoje.replace(day=1)
    fim_mes = (inicio_mes + relativedelta(months=1)) - timedelta(days=1)

    # Totais do mês atual
    entradas_mes = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'entrada',
        Transaction.date >= inicio_mes,
        Transaction.date <= fim_mes,
        Transaction.status == 'realizado'
    ).scalar() or 0

    saidas_mes = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio_mes,
        Transaction.date <= fim_mes,
        Transaction.status == 'realizado'
    ).scalar() or 0

    saldo_mes = entradas_mes - saidas_mes

    # Saldo acumulado total
    total_entradas = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'entrada',
        Transaction.status == 'realizado'
    ).scalar() or 0
    total_saidas = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'saida',
        Transaction.status == 'realizado'
    ).scalar() or 0
    saldo_total = total_entradas - total_saidas

    # Últimos 6 meses para gráfico
    meses_labels = []
    meses_entradas = []
    meses_saidas = []
    for i in range(5, -1, -1):
        mes_ref = hoje - relativedelta(months=i)
        inicio = mes_ref.replace(day=1)
        fim = (inicio + relativedelta(months=1)) - timedelta(days=1)
        e = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'entrada',
            Transaction.date >= inicio,
            Transaction.date <= fim,
            Transaction.status == 'realizado'
        ).scalar() or 0
        s = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'saida',
            Transaction.date >= inicio,
            Transaction.date <= fim,
            Transaction.status == 'realizado'
        ).scalar() or 0
        meses_labels.append(mes_ref.strftime('%b/%Y'))
        meses_entradas.append(float(e))
        meses_saidas.append(float(s))

    # Lançamentos recentes
    lancamentos_recentes = Transaction.query.order_by(
        Transaction.date.desc(), Transaction.created_at.desc()
    ).limit(8).all()

    # Contratos ativos
    contratos_ativos = Contract.query.filter_by(status='ativo').count()

    # Previsão próximos 30 dias (projeções)
    hoje_dt = hoje
    fim_30 = hoje_dt + timedelta(days=30)
    prev_entradas = db.session.query(func.sum(Projection.value)).filter(
        Projection.type == 'entrada',
        Projection.date >= hoje_dt,
        Projection.date <= fim_30
    ).scalar() or 0
    prev_saidas = db.session.query(func.sum(Projection.value)).filter(
        Projection.type == 'saida',
        Projection.date >= hoje_dt,
        Projection.date <= fim_30
    ).scalar() or 0

    # Gráfico pizza - saídas por categoria no mês
    pizza_saidas = db.session.query(
        Category.name, func.sum(Transaction.value).label('total')
    ).join(Transaction, Transaction.category_id == Category.id).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio_mes,
        Transaction.date <= fim_mes,
        Transaction.status == 'realizado'
    ).group_by(Category.id).order_by(func.sum(Transaction.value).desc()).limit(8).all()

    pizza_labels = [r[0] for r in pizza_saidas]
    pizza_values = [float(r[1]) for r in pizza_saidas]

    # Alerta de saldo baixo
    limite_alerta = float(os.environ.get('ALERTA_SALDO', 10000))
    alerta_saldo = saldo_total < limite_alerta

    # Lembretes no dashboard
    BillReminder.query.filter(
        BillReminder.status == 'pendente',
        BillReminder.due_date < hoje
    ).update({'status': 'atrasado'})
    db.session.commit()

    lembretes_atrasados = BillReminder.query.filter_by(status='atrasado').order_by(BillReminder.due_date).limit(5).all()
    lembretes_proximos = BillReminder.query.filter(
        BillReminder.status == 'pendente',
        BillReminder.due_date <= hoje + timedelta(days=7)
    ).order_by(BillReminder.due_date).limit(5).all()
    total_atrasados = BillReminder.query.filter_by(status='atrasado').count()
    total_proximos = BillReminder.query.filter(
        BillReminder.status == 'pendente',
        BillReminder.due_date <= hoje + timedelta(days=7)
    ).count()

    return render_template('dashboard.html',
        entradas_mes=entradas_mes, saidas_mes=saidas_mes,
        saldo_mes=saldo_mes, saldo_total=saldo_total,
        meses_labels=json.dumps(meses_labels),
        meses_entradas=json.dumps(meses_entradas),
        meses_saidas=json.dumps(meses_saidas),
        lancamentos_recentes=lancamentos_recentes,
        contratos_ativos=contratos_ativos,
        prev_entradas=prev_entradas, prev_saidas=prev_saidas,
        pizza_labels=json.dumps(pizza_labels),
        pizza_values=json.dumps(pizza_values),
        alerta_saldo=alerta_saldo,
        limite_alerta=limite_alerta,
        lembretes_atrasados=lembretes_atrasados,
        lembretes_proximos=lembretes_proximos,
        total_atrasados=total_atrasados,
        total_proximos=total_proximos,
        hoje=hoje
    )

# ─── LANÇAMENTOS ─────────────────────────────────────────────────────────────

@app.route('/lancamentos')
@login_required
def lancamentos():
    page = request.args.get('page', 1, type=int)
    tipo = request.args.get('tipo', '')
    status = request.args.get('status', '')
    data_ini = request.args.get('data_ini', '')
    data_fim = request.args.get('data_fim', '')
    categoria_id = request.args.get('categoria_id', '', type=str)

    q = Transaction.query
    if tipo:
        q = q.filter(Transaction.type == tipo)
    if status:
        q = q.filter(Transaction.status == status)
    if data_ini:
        q = q.filter(Transaction.date >= datetime.strptime(data_ini, '%Y-%m-%d').date())
    if data_fim:
        q = q.filter(Transaction.date <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    if categoria_id:
        q = q.filter(Transaction.category_id == int(categoria_id))

    lancamentos_paginados = q.order_by(Transaction.date.desc(), Transaction.created_at.desc()).paginate(page=page, per_page=15)
    categorias = Category.query.filter_by(active=True).order_by(Category.name).all()

    return render_template('lancamentos/index.html',
        lancamentos=lancamentos_paginados,
        categorias=categorias,
        filtros={'tipo': tipo, 'status': status, 'data_ini': data_ini, 'data_fim': data_fim, 'categoria_id': categoria_id}
    )

@app.route('/lancamentos/novo', methods=['GET', 'POST'])
@login_required
def lancamento_novo():
    categorias = Category.query.filter_by(active=True).order_by(Category.type, Category.name).all()
    contratos = Contract.query.filter_by(status='ativo').order_by(Contract.number).all()
    centros = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    if request.method == 'POST':
        try:
            valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
            t = Transaction(
                date=datetime.strptime(request.form['date'], '%Y-%m-%d').date(),
                description=request.form['description'].strip(),
                category_id=int(request.form['category_id']),
                contract_id=int(request.form['contract_id']) if request.form.get('contract_id') else None,
                cost_center_id=int(request.form['cost_center_id']) if request.form.get('cost_center_id') else None,
                value=float(valor_str),
                type=request.form['type'],
                status=request.form['status'],
                notes=request.form.get('notes', '').strip(),
                user_id=current_user.id
            )
            arquivo = request.files.get('attachment')
            if arquivo and arquivo.filename:
                if not allowed_file(arquivo.filename):
                    flash('Formato de arquivo não permitido. Use PDF, JPG, PNG ou GIF.', 'danger')
                    return render_template('lancamentos/form.html', lancamento=None,
                                           categorias=categorias, contratos=contratos,
                                           centros=centros, hoje=date.today())
                t.attachment_data = arquivo.read()
                t.attachment_original = arquivo.filename
                t.attachment_mimetype = arquivo.mimetype or mimetypes.guess_type(arquivo.filename)[0] or 'application/octet-stream'
            db.session.add(t)
            db.session.commit()
            flash('Lançamento registrado com sucesso!', 'success')
            return redirect(url_for('lancamentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {str(e)}', 'danger')
    return render_template('lancamentos/form.html', lancamento=None, categorias=categorias,
                           contratos=contratos, centros=centros, hoje=date.today())

@app.route('/lancamentos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def lancamento_editar(id):
    t = Transaction.query.get_or_404(id)
    categorias = Category.query.filter_by(active=True).order_by(Category.type, Category.name).all()
    contratos = Contract.query.filter_by(status='ativo').order_by(Contract.number).all()
    centros = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    if request.method == 'POST':
        try:
            valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
            t.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
            t.description = request.form['description'].strip()
            t.category_id = int(request.form['category_id'])
            t.contract_id = int(request.form['contract_id']) if request.form.get('contract_id') else None
            t.cost_center_id = int(request.form['cost_center_id']) if request.form.get('cost_center_id') else None
            t.value = float(valor_str)
            t.type = request.form['type']
            t.status = request.form['status']
            t.notes = request.form.get('notes', '').strip()
            arquivo = request.files.get('attachment')
            if arquivo and arquivo.filename:
                if not allowed_file(arquivo.filename):
                    flash('Formato de arquivo não permitido. Use PDF, JPG, PNG ou GIF.', 'danger')
                    return render_template('lancamentos/form.html', lancamento=t,
                                           categorias=categorias, contratos=contratos, centros=centros)
                t.attachment_data = arquivo.read()
                t.attachment_original = arquivo.filename
                t.attachment_mimetype = arquivo.mimetype or mimetypes.guess_type(arquivo.filename)[0] or 'application/octet-stream'
            db.session.commit()
            flash('Lançamento atualizado com sucesso!', 'success')
            return redirect(url_for('lancamentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return render_template('lancamentos/form.html', lancamento=t, categorias=categorias,
                           contratos=contratos, centros=centros)

@app.route('/lancamentos/<int:id>/anexo')
@login_required
def lancamento_anexo(id):
    t = Transaction.query.get_or_404(id)
    if not t.attachment_data:
        flash('Este lançamento não possui anexo.', 'warning')
        return redirect(url_for('lancamentos'))
    return send_file(
        io.BytesIO(t.attachment_data),
        mimetype=t.attachment_mimetype or 'application/octet-stream',
        as_attachment=False,
        download_name=t.attachment_original or 'anexo'
    )

@app.route('/lancamentos/<int:id>/anexo/download')
@login_required
def lancamento_anexo_download(id):
    t = Transaction.query.get_or_404(id)
    if not t.attachment_data:
        flash('Este lançamento não possui anexo.', 'warning')
        return redirect(url_for('lancamentos'))
    return send_file(
        io.BytesIO(t.attachment_data),
        mimetype=t.attachment_mimetype or 'application/octet-stream',
        as_attachment=True,
        download_name=t.attachment_original or 'anexo'
    )

@app.route('/lancamentos/<int:id>/anexo/excluir', methods=['POST'])
@login_required
def lancamento_anexo_excluir(id):
    t = Transaction.query.get_or_404(id)
    if not current_user.can_edit():
        flash('Sem permissão.', 'danger')
        return redirect(url_for('lancamento_editar', id=id))
    t.attachment_data = None
    t.attachment_original = None
    t.attachment_mimetype = None
    db.session.commit()
    flash('Anexo removido.', 'success')
    return redirect(url_for('lancamento_editar', id=id))

@app.route('/lancamentos/<int:id>/excluir', methods=['POST'])
@login_required
def lancamento_excluir(id):
    t = Transaction.query.get_or_404(id)
    db.session.delete(t)
    db.session.commit()
    flash('Lançamento excluído.', 'success')
    return redirect(url_for('lancamentos'))

# ─── CONTRATOS ───────────────────────────────────────────────────────────────

@app.route('/contratos')
@login_required
def contratos():
    status = request.args.get('status', '')
    q = Contract.query
    if status:
        q = q.filter(Contract.status == status)
    contratos = q.order_by(Contract.created_at.desc()).all()
    return render_template('contratos/index.html', contratos=contratos, filtro_status=status)

@app.route('/contratos/novo', methods=['GET', 'POST'])
@login_required
def contrato_novo():
    if request.method == 'POST':
        try:
            valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
            c = Contract(
                number=request.form['number'].strip(),
                client=request.form['client'].strip(),
                description=request.form['description'].strip(),
                value=float(valor_str),
                start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
                end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None,
                status=request.form.get('status', 'ativo')
            )
            db.session.add(c)
            db.session.commit()
            flash('Contrato cadastrado com sucesso!', 'success')
            return redirect(url_for('contratos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {str(e)}', 'danger')
    return render_template('contratos/form.html', contrato=None)

@app.route('/contratos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def contrato_editar(id):
    c = Contract.query.get_or_404(id)
    if request.method == 'POST':
        try:
            valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
            c.number = request.form['number'].strip()
            c.client = request.form['client'].strip()
            c.description = request.form['description'].strip()
            c.value = float(valor_str)
            c.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
            c.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None
            c.status = request.form.get('status', 'ativo')
            db.session.commit()
            flash('Contrato atualizado com sucesso!', 'success')
            return redirect(url_for('contratos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return render_template('contratos/form.html', contrato=c)

@app.route('/contratos/<int:id>/excluir', methods=['POST'])
@login_required
def contrato_excluir(id):
    c = Contract.query.get_or_404(id)
    db.session.delete(c)
    db.session.commit()
    flash('Contrato excluído.', 'success')
    return redirect(url_for('contratos'))

# ─── PROJEÇÕES ────────────────────────────────────────────────────────────────

@app.route('/projecoes')
@login_required
def projecoes():
    hoje = date.today()
    data_ini = request.args.get('data_ini', hoje.strftime('%Y-%m-%d'))
    data_fim = request.args.get('data_fim', (hoje + timedelta(days=90)).strftime('%Y-%m-%d'))
    tipo = request.args.get('tipo', '')

    q = Projection.query
    if data_ini:
        q = q.filter(Projection.date >= datetime.strptime(data_ini, '%Y-%m-%d').date())
    if data_fim:
        q = q.filter(Projection.date <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    if tipo:
        q = q.filter(Projection.type == tipo)

    projecoes = q.order_by(Projection.date.asc()).all()

    # Acumular saldo projetado
    saldo_atual = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'entrada', Transaction.status == 'realizado'
    ).scalar() or 0
    saldo_atual -= db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'saida', Transaction.status == 'realizado'
    ).scalar() or 0

    # Dados do gráfico de projeção (agrupado por semana)
    proj_labels = []
    proj_saldos = []
    saldo_acum = float(saldo_atual)
    if projecoes:
        from itertools import groupby
        for p in projecoes:
            if p.type == 'entrada':
                saldo_acum += float(p.value)
            else:
                saldo_acum -= float(p.value)
            proj_labels.append(p.date.strftime('%d/%m'))
            proj_saldos.append(round(saldo_acum, 2))

    categorias = Category.query.filter_by(active=True).order_by(Category.name).all()
    contratos = Contract.query.filter_by(status='ativo').all()

    return render_template('projecoes/index.html',
        projecoes=projecoes, saldo_atual=saldo_atual,
        proj_labels=json.dumps(proj_labels), proj_saldos=json.dumps(proj_saldos),
        categorias=categorias, contratos=contratos,
        filtros={'data_ini': data_ini, 'data_fim': data_fim, 'tipo': tipo}
    )

@app.route('/projecoes/novo', methods=['POST'])
@login_required
def projecao_nova():
    try:
        valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
        p = Projection(
            date=datetime.strptime(request.form['date'], '%Y-%m-%d').date(),
            description=request.form['description'].strip(),
            category_id=int(request.form['category_id']),
            contract_id=int(request.form['contract_id']) if request.form.get('contract_id') else None,
            value=float(valor_str),
            type=request.form['type'],
            notes=request.form.get('notes', '').strip()
        )
        db.session.add(p)
        db.session.commit()
        flash('Projeção adicionada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar: {str(e)}', 'danger')
    return redirect(url_for('projecoes'))

@app.route('/projecoes/<int:id>/excluir', methods=['POST'])
@login_required
def projecao_excluir(id):
    p = Projection.query.get_or_404(id)
    db.session.delete(p)
    db.session.commit()
    flash('Projeção excluída.', 'success')
    return redirect(url_for('projecoes'))

@app.route('/projecoes/<int:id>/realizar', methods=['POST'])
@login_required
def projecao_realizar(id):
    p = Projection.query.get_or_404(id)
    t = Transaction(
        date=p.date,
        description=p.description,
        category_id=p.category_id,
        contract_id=p.contract_id,
        value=p.value,
        type=p.type,
        status='realizado',
        notes=f'Realizado a partir da projeção. {p.notes}',
        user_id=current_user.id
    )
    db.session.add(t)
    db.session.delete(p)
    db.session.commit()
    flash('Projeção lançada como realizada!', 'success')
    return redirect(url_for('projecoes'))

# ─── RELATÓRIOS ───────────────────────────────────────────────────────────────

@app.route('/relatorios')
@login_required
def relatorios():
    hoje = date.today()
    mes = request.args.get('mes', hoje.month, type=int)
    ano = request.args.get('ano', hoje.year, type=int)

    inicio = date(ano, mes, 1)
    fim = (inicio + relativedelta(months=1)) - timedelta(days=1)

    # Totais
    entradas = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'entrada',
        Transaction.date >= inicio,
        Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).scalar() or 0

    saidas = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio,
        Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).scalar() or 0

    saldo = entradas - saidas

    # Por categoria - entradas
    cat_entradas = db.session.query(
        Category.name, func.sum(Transaction.value).label('total')
    ).join(Transaction, Transaction.category_id == Category.id).filter(
        Transaction.type == 'entrada',
        Transaction.date >= inicio,
        Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).group_by(Category.id).order_by(func.sum(Transaction.value).desc()).all()

    cat_saidas = db.session.query(
        Category.name, func.sum(Transaction.value).label('total')
    ).join(Transaction, Transaction.category_id == Category.id).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio,
        Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).group_by(Category.id).order_by(func.sum(Transaction.value).desc()).all()

    # Todos os lançamentos do período
    lancamentos = Transaction.query.filter(
        Transaction.date >= inicio,
        Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).order_by(Transaction.date.asc(), Transaction.type.asc()).all()

    # Dados para gráfico diário
    dias_labels = []
    dias_entradas_vals = []
    dias_saidas_vals = []
    d = inicio
    while d <= fim:
        e = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'entrada',
            Transaction.date == d,
            Transaction.status == 'realizado'
        ).scalar() or 0
        s = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'saida',
            Transaction.date == d,
            Transaction.status == 'realizado'
        ).scalar() or 0
        dias_labels.append(d.strftime('%d'))
        dias_entradas_vals.append(float(e))
        dias_saidas_vals.append(float(s))
        d += timedelta(days=1)

    meses_disponiveis = []
    for y in range(hoje.year - 2, hoje.year + 1):
        for m in range(1, 13):
            meses_disponiveis.append({'ano': y, 'mes': m,
                'label': date(y, m, 1).strftime('%B/%Y').capitalize()})

    return render_template('relatorios/index.html',
        entradas=entradas, saidas=saidas, saldo=saldo,
        cat_entradas=cat_entradas, cat_saidas=cat_saidas,
        lancamentos=lancamentos, inicio=inicio, fim=fim,
        mes=mes, ano=ano,
        dias_labels=json.dumps(dias_labels),
        dias_entradas_vals=json.dumps(dias_entradas_vals),
        dias_saidas_vals=json.dumps(dias_saidas_vals),
        meses_disponiveis=meses_disponiveis
    )

@app.route('/relatorios/pdf')
@login_required
def relatorio_pdf():
    mes = request.args.get('mes', date.today().month, type=int)
    ano = request.args.get('ano', date.today().year, type=int)
    inicio = date(ano, mes, 1)
    fim = (inicio + relativedelta(months=1)) - timedelta(days=1)

    entradas_total = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'entrada',
        Transaction.date >= inicio, Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).scalar() or 0

    saidas_total = db.session.query(func.sum(Transaction.value)).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio, Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).scalar() or 0

    lancamentos = Transaction.query.filter(
        Transaction.date >= inicio, Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).order_by(Transaction.date.asc(), Transaction.type.asc()).all()

    cat_entradas = db.session.query(
        Category.name, func.sum(Transaction.value).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'entrada',
        Transaction.date >= inicio, Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).group_by(Category.id).order_by(func.sum(Transaction.value).desc()).all()

    cat_saidas = db.session.query(
        Category.name, func.sum(Transaction.value).label('total')
    ).join(Transaction).filter(
        Transaction.type == 'saida',
        Transaction.date >= inicio, Transaction.date <= fim,
        Transaction.status == 'realizado'
    ).group_by(Category.id).order_by(func.sum(Transaction.value).desc()).all()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Cabeçalho
    pdf.set_fill_color(30, 64, 175)
    pdf.rect(0, 0, 210, 35, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 18)
    pdf.set_xy(10, 8)
    pdf.cell(190, 10, 'DEMONSTRACAO DE FLUXO DE CAIXA', align='C')
    pdf.set_font('Helvetica', '', 11)
    pdf.set_xy(10, 20)
    pdf.cell(190, 8, f'Periodo: {inicio.strftime("%d/%m/%Y")} a {fim.strftime("%d/%m/%Y")}', align='C')
    pdf.set_xy(10, 27)
    pdf.cell(190, 6, f'Emitido em: {datetime.now().strftime("%d/%m/%Y as %H:%M")}', align='C')

    pdf.set_text_color(0, 0, 0)
    pdf.set_y(42)

    # Resumo
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_fill_color(243, 244, 246)
    pdf.cell(190, 8, 'RESUMO DO PERIODO', fill=True, ln=True)
    pdf.ln(2)

    def fmt_valor(v):
        return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    pdf.set_font('Helvetica', '', 11)
    pdf.set_fill_color(220, 252, 231)
    pdf.cell(95, 8, 'Total de Entradas', fill=True, border=1)
    pdf.set_fill_color(220, 252, 231)
    pdf.cell(95, 8, fmt_valor(entradas_total), fill=True, border=1, align='R', ln=True)

    pdf.set_fill_color(254, 226, 226)
    pdf.cell(95, 8, 'Total de Saidas', fill=True, border=1)
    pdf.cell(95, 8, fmt_valor(saidas_total), fill=True, border=1, align='R', ln=True)

    saldo = entradas_total - saidas_total
    fill_cor = (220, 252, 231) if saldo >= 0 else (254, 226, 226)
    pdf.set_fill_color(*fill_cor)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(95, 9, 'SALDO DO PERIODO', fill=True, border=1)
    pdf.cell(95, 9, fmt_valor(saldo), fill=True, border=1, align='R', ln=True)
    pdf.ln(4)

    # Por categoria - Entradas
    if cat_entradas:
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_fill_color(243, 244, 246)
        pdf.cell(190, 8, 'ENTRADAS POR CATEGORIA', fill=True, ln=True)
        pdf.ln(1)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_fill_color(209, 250, 229)
        pdf.cell(130, 7, 'Categoria', fill=True, border=1)
        pdf.cell(60, 7, 'Valor', fill=True, border=1, align='R', ln=True)
        pdf.set_font('Helvetica', '', 10)
        for cat, total in cat_entradas:
            pdf.cell(130, 7, str(cat), border=1)
            pdf.cell(60, 7, fmt_valor(total), border=1, align='R', ln=True)
        pdf.ln(4)

    # Por categoria - Saídas
    if cat_saidas:
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_fill_color(243, 244, 246)
        pdf.cell(190, 8, 'SAIDAS POR CATEGORIA', fill=True, ln=True)
        pdf.ln(1)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_fill_color(254, 205, 211)
        pdf.cell(130, 7, 'Categoria', fill=True, border=1)
        pdf.cell(60, 7, 'Valor', fill=True, border=1, align='R', ln=True)
        pdf.set_font('Helvetica', '', 10)
        for cat, total in cat_saidas:
            pdf.cell(130, 7, str(cat), border=1)
            pdf.cell(60, 7, fmt_valor(total), border=1, align='R', ln=True)
        pdf.ln(4)

    # Lançamentos detalhados
    if lancamentos:
        pdf.add_page()
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_fill_color(243, 244, 246)
        pdf.cell(190, 8, 'LANCAMENTOS DETALHADOS', fill=True, ln=True)
        pdf.ln(1)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(224, 231, 255)
        pdf.cell(22, 7, 'Data', fill=True, border=1)
        pdf.cell(75, 7, 'Descricao', fill=True, border=1)
        pdf.cell(40, 7, 'Categoria', fill=True, border=1)
        pdf.cell(18, 7, 'Tipo', fill=True, border=1, align='C')
        pdf.cell(35, 7, 'Valor', fill=True, border=1, align='R', ln=True)
        pdf.set_font('Helvetica', '', 9)
        for l in lancamentos:
            tipo_txt = 'Entrada' if l.type == 'entrada' else 'Saida'
            desc = l.description[:45] if len(l.description) > 45 else l.description
            cat_nome = l.category.name[:22] if l.category else '-'
            pdf.cell(22, 6, l.date.strftime('%d/%m/%Y'), border=1)
            pdf.cell(75, 6, desc, border=1)
            pdf.cell(40, 6, cat_nome, border=1)
            pdf.cell(18, 6, tipo_txt, border=1, align='C')
            pdf.cell(35, 6, fmt_valor(l.value), border=1, align='R', ln=True)

    pdf_bytes = pdf.output()
    response = make_response(bytes(pdf_bytes))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=DFC_{ano}_{mes:02d}.pdf'
    return response

# ─── RELATÓRIO COMPARATIVO ────────────────────────────────────────────────────

@app.route('/relatorios/comparativo')
@login_required
def relatorio_comparativo():
    hoje = date.today()
    mes1 = request.args.get('mes1', hoje.month, type=int)
    ano1 = request.args.get('ano1', hoje.year, type=int)
    mes2_default = (hoje - relativedelta(months=1)).month
    ano2_default = (hoje - relativedelta(months=1)).year
    mes2 = request.args.get('mes2', mes2_default, type=int)
    ano2 = request.args.get('ano2', ano2_default, type=int)

    def get_totais(ano, mes):
        inicio = date(ano, mes, 1)
        fim = (inicio + relativedelta(months=1)) - timedelta(days=1)
        e = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'entrada', Transaction.date >= inicio,
            Transaction.date <= fim, Transaction.status == 'realizado'
        ).scalar() or 0
        s = db.session.query(func.sum(Transaction.value)).filter(
            Transaction.type == 'saida', Transaction.date >= inicio,
            Transaction.date <= fim, Transaction.status == 'realizado'
        ).scalar() or 0
        cats = db.session.query(
            Category.name, Category.type, func.sum(Transaction.value).label('total')
        ).join(Transaction, Transaction.category_id == Category.id).filter(
            Transaction.date >= inicio, Transaction.date <= fim,
            Transaction.status == 'realizado'
        ).group_by(Category.id).order_by(Category.type, func.sum(Transaction.value).desc()).all()
        return {'entradas': float(e), 'saidas': float(s), 'saldo': float(e - s), 'cats': cats}

    d1 = get_totais(ano1, mes1)
    d2 = get_totais(ano2, mes2)

    meses_disponiveis = []
    for y in range(hoje.year - 2, hoje.year + 1):
        for m in range(1, 13):
            meses_disponiveis.append({'ano': y, 'mes': m,
                'label': date(y, m, 1).strftime('%B/%Y').capitalize()})

    return render_template('relatorios/comparativo.html',
        d1=d1, d2=d2,
        mes1=mes1, ano1=ano1, mes2=mes2, ano2=ano2,
        label1=date(ano1, mes1, 1).strftime('%B/%Y').capitalize(),
        label2=date(ano2, mes2, 1).strftime('%B/%Y').capitalize(),
        meses_disponiveis=meses_disponiveis
    )

# ─── EXPORTAÇÃO EXCEL ─────────────────────────────────────────────────────────

@app.route('/lancamentos/exportar-excel')
@login_required
def lancamentos_exportar_excel():
    tipo = request.args.get('tipo', '')
    status = request.args.get('status', '')
    data_ini = request.args.get('data_ini', '')
    data_fim = request.args.get('data_fim', '')
    categoria_id = request.args.get('categoria_id', '')

    q = Transaction.query
    if tipo:
        q = q.filter(Transaction.type == tipo)
    if status:
        q = q.filter(Transaction.status == status)
    if data_ini:
        q = q.filter(Transaction.date >= datetime.strptime(data_ini, '%Y-%m-%d').date())
    if data_fim:
        q = q.filter(Transaction.date <= datetime.strptime(data_fim, '%Y-%m-%d').date())
    if categoria_id:
        q = q.filter(Transaction.category_id == int(categoria_id))
    lancamentos = q.order_by(Transaction.date.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lançamentos'

    header_fill = PatternFill(start_color='1e2235', end_color='1e2235', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )

    headers = ['Data', 'Descrição', 'Categoria', 'Contrato', 'Tipo', 'Valor (R$)', 'Status', 'Observações']
    col_widths = [14, 40, 25, 20, 10, 18, 14, 40]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    green_fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
    red_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')

    for row_num, t in enumerate(lancamentos, 2):
        row_fill = green_fill if t.type == 'entrada' else red_fill
        data = [
            t.date.strftime('%d/%m/%Y'),
            t.description,
            t.category.name if t.category else '-',
            t.contract.number if t.contract else '-',
            'Entrada' if t.type == 'entrada' else 'Saída',
            float(t.value),
            t.status.capitalize(),
            t.notes or ''
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col_num == 6:
                cell.number_format = '#,##0.00'
    ws.row_dimensions[1].height = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'lancamentos_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ─── IMPORTAÇÃO CSV/EXCEL ─────────────────────────────────────────────────────

@app.route('/lancamentos/importar', methods=['GET', 'POST'])
@login_required
def lancamentos_importar():
    if not current_user.can_edit():
        flash('Sem permissão para importar lançamentos.', 'danger')
        return redirect(url_for('lancamentos'))

    categorias = Category.query.filter_by(active=True).order_by(Category.type, Category.name).all()
    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if not arquivo or arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'danger')
            return redirect(request.url)

        filename = arquivo.filename.lower()
        erros = []
        importados = 0
        try:
            if filename.endswith('.csv'):
                content = arquivo.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(arquivo, data_only=True) if filename.endswith('.xlsx') else None
                if wb is None:
                    import xlrd
                    book = xlrd.open_workbook(file_contents=arquivo.read())
                    sheet = book.sheet_by_index(0)
                    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
                    rows = [dict(zip(headers, [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)])) for r in range(1, sheet.nrows)]
                else:
                    ws = wb.active
                    headers = [str(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
                    rows = []
                    for r in range(2, ws.max_row + 1):
                        rows.append({headers[c-1]: str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)})
            else:
                flash('Formato inválido. Use CSV, XLSX ou XLS.', 'danger')
                return redirect(request.url)

            cat_map = {c.name.lower(): c for c in categorias}
            for i, row in enumerate(rows, 2):
                try:
                    data_str = row.get('data', row.get('Data', '')).strip()
                    desc = row.get('descricao', row.get('descrição', row.get('Descricao', row.get('Descrição', '')))).strip()
                    tipo = row.get('tipo', row.get('Tipo', '')).strip().lower()
                    valor_str = str(row.get('valor', row.get('Valor', '0'))).replace('R$', '').replace('.', '').replace(',', '.').strip()
                    cat_nome = row.get('categoria', row.get('Categoria', '')).strip().lower()
                    status_imp = row.get('status', row.get('Status', 'realizado')).strip().lower()

                    if not data_str or not desc or not tipo or not valor_str:
                        erros.append(f'Linha {i}: campos obrigatórios ausentes.')
                        continue
                    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                        try:
                            dt = datetime.strptime(data_str, fmt).date()
                            break
                        except ValueError:
                            dt = None
                    if not dt:
                        erros.append(f'Linha {i}: data inválida "{data_str}".')
                        continue
                    if tipo not in ('entrada', 'saida', 'saída'):
                        erros.append(f'Linha {i}: tipo deve ser "entrada" ou "saida".')
                        continue
                    tipo_norm = 'saida' if 'sa' in tipo else 'entrada'
                    valor = float(valor_str)
                    cat = cat_map.get(cat_nome)
                    if not cat:
                        cat = Category.query.filter_by(type=tipo_norm, active=True).first()
                    status_norm = 'realizado' if 'realiz' in status_imp else 'previsto'

                    db.session.add(Transaction(
                        date=dt, description=desc, category_id=cat.id if cat else None,
                        value=valor, type=tipo_norm, status=status_norm,
                        user_id=current_user.id
                    ))
                    importados += 1
                except Exception as e:
                    erros.append(f'Linha {i}: {str(e)}')

            db.session.commit()
            msg = f'{importados} lançamento(s) importado(s) com sucesso.'
            if erros:
                msg += f' {len(erros)} linha(s) com erro ignorada(s).'
            flash(msg, 'success' if importados > 0 else 'warning')
            if erros:
                for e in erros[:5]:
                    flash(e, 'warning')
            return redirect(url_for('lancamentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')

    return render_template('lancamentos/importar.html', categorias=categorias)

# ─── CATEGORIAS ───────────────────────────────────────────────────────────────

@app.route('/categorias')
@login_required
def categorias():
    cats = Category.query.order_by(Category.type, Category.name).all()
    return render_template('categorias/index.html', categorias=cats)

@app.route('/categorias/nova', methods=['POST'])
@login_required
def categoria_nova():
    name = request.form.get('name', '').strip()
    tipo = request.form.get('type', 'saida')
    if name:
        c = Category(name=name, type=tipo, active=True)
        db.session.add(c)
        db.session.commit()
        flash('Categoria criada com sucesso!', 'success')
    return redirect(url_for('categorias'))

@app.route('/categorias/<int:id>/toggle', methods=['POST'])
@login_required
def categoria_toggle(id):
    c = Category.query.get_or_404(id)
    c.active = not c.active
    db.session.commit()
    return redirect(url_for('categorias'))

@app.route('/categorias/<int:id>/excluir', methods=['POST'])
@login_required
def categoria_excluir(id):
    c = Category.query.get_or_404(id)
    if c.transactions:
        flash('Não é possível excluir categoria com lançamentos vinculados.', 'warning')
    else:
        db.session.delete(c)
        db.session.commit()
        flash('Categoria excluída.', 'success')
    return redirect(url_for('categorias'))

# ─── API ─────────────────────────────────────────────────────────────────────

@app.route('/api/categorias/<tipo>')
@login_required
def api_categorias(tipo):
    cats = Category.query.filter_by(type=tipo, active=True).order_by(Category.name).all()
    return jsonify([{'id': c.id, 'name': c.name} for c in cats])

@app.route('/api/centros-custo')
@login_required
def api_centros_custo():
    ccs = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    return jsonify([{'id': c.id, 'name': c.name, 'code': c.code or ''} for c in ccs])

# ─── CENTRO DE CUSTO ──────────────────────────────────────────────────────────

@app.route('/centro-custo')
@login_required
def centro_custo():
    ccs = CostCenter.query.order_by(CostCenter.status.desc(), CostCenter.name).all()
    return render_template('centro_custo/index.html', centros=ccs)

@app.route('/centro-custo/novo', methods=['GET', 'POST'])
@login_required
def centro_custo_novo():
    if not current_user.can_edit():
        flash('Sem permissão.', 'danger')
        return redirect(url_for('centro_custo'))
    contratos = Contract.query.filter_by(status='ativo').order_by(Contract.number).all()
    if request.method == 'POST':
        cc = CostCenter(
            code=request.form.get('code', '').strip() or None,
            name=request.form['name'].strip(),
            description=request.form.get('description', '').strip(),
            contract_id=int(request.form['contract_id']) if request.form.get('contract_id') else None,
            status='ativo'
        )
        db.session.add(cc)
        db.session.commit()
        flash('Centro de custo criado!', 'success')
        return redirect(url_for('centro_custo'))
    return render_template('centro_custo/form.html', cc=None, contratos=contratos)

@app.route('/centro-custo/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def centro_custo_editar(id):
    if not current_user.can_edit():
        flash('Sem permissão.', 'danger')
        return redirect(url_for('centro_custo'))
    cc = CostCenter.query.get_or_404(id)
    contratos = Contract.query.order_by(Contract.number).all()
    if request.method == 'POST':
        cc.code = request.form.get('code', '').strip() or None
        cc.name = request.form['name'].strip()
        cc.description = request.form.get('description', '').strip()
        cc.contract_id = int(request.form['contract_id']) if request.form.get('contract_id') else None
        cc.status = request.form.get('status', 'ativo')
        db.session.commit()
        flash('Centro de custo atualizado!', 'success')
        return redirect(url_for('centro_custo'))
    return render_template('centro_custo/form.html', cc=cc, contratos=contratos)

@app.route('/centro-custo/<int:id>/toggle', methods=['POST'])
@login_required
def centro_custo_toggle(id):
    cc = CostCenter.query.get_or_404(id)
    cc.status = 'inativo' if cc.status == 'ativo' else 'ativo'
    db.session.commit()
    return redirect(url_for('centro_custo'))

@app.route('/centro-custo/<int:id>/excluir', methods=['POST'])
@login_required
def centro_custo_excluir(id):
    cc = CostCenter.query.get_or_404(id)
    if cc.transactions or cc.bill_reminders:
        flash('Não é possível excluir: existem lançamentos ou lembretes vinculados.', 'warning')
    else:
        db.session.delete(cc)
        db.session.commit()
        flash('Centro de custo excluído.', 'success')
    return redirect(url_for('centro_custo'))

# ─── LEMBRETES DE CONTAS A PAGAR ──────────────────────────────────────────────

@app.route('/lembretes')
@login_required
def lembretes():
    hoje = date.today()
    # Atualiza status de pendentes atrasados automaticamente
    BillReminder.query.filter(
        BillReminder.status == 'pendente',
        BillReminder.due_date < hoje
    ).update({'status': 'atrasado'})
    db.session.commit()

    status_filtro = request.args.get('status', '')
    q = BillReminder.query
    if status_filtro:
        q = q.filter(BillReminder.status == status_filtro)
    lembretes = q.order_by(BillReminder.due_date.asc()).all()

    # Contadores para badges
    atrasados = BillReminder.query.filter_by(status='atrasado').count()
    proximos  = BillReminder.query.filter(
        BillReminder.status == 'pendente',
        BillReminder.due_date <= hoje + timedelta(days=7)
    ).count()

    categorias = Category.query.filter_by(type='saida', active=True).order_by(Category.name).all()
    centros = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    return render_template('lembretes/index.html',
        lembretes=lembretes, hoje=hoje,
        atrasados=atrasados, proximos=proximos,
        status_filtro=status_filtro,
        categorias=categorias, centros=centros)

@app.route('/lembretes/novo', methods=['GET', 'POST'])
@login_required
def lembrete_novo():
    categorias = Category.query.filter_by(type='saida', active=True).order_by(Category.name).all()
    centros = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    if request.method == 'POST':
        valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
        b = BillReminder(
            description=request.form['description'].strip(),
            value=float(valor_str),
            due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d').date(),
            category_id=int(request.form['category_id']) if request.form.get('category_id') else None,
            cost_center_id=int(request.form['cost_center_id']) if request.form.get('cost_center_id') else None,
            recurrence=request.form.get('recurrence', 'nenhuma'),
            status='pendente',
            notes=request.form.get('notes', '').strip(),
            user_id=current_user.id
        )
        db.session.add(b)
        db.session.commit()
        flash('Lembrete criado!', 'success')
        return redirect(url_for('lembretes'))
    return render_template('lembretes/form.html', lembrete=None,
                           categorias=categorias, centros=centros, hoje=date.today())

@app.route('/lembretes/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def lembrete_editar(id):
    b = BillReminder.query.get_or_404(id)
    categorias = Category.query.filter_by(type='saida', active=True).order_by(Category.name).all()
    centros = CostCenter.query.filter_by(status='ativo').order_by(CostCenter.name).all()
    if request.method == 'POST':
        valor_str = request.form.get('value', '0').replace('.', '').replace(',', '.')
        b.description = request.form['description'].strip()
        b.value = float(valor_str)
        b.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d').date()
        b.category_id = int(request.form['category_id']) if request.form.get('category_id') else None
        b.cost_center_id = int(request.form['cost_center_id']) if request.form.get('cost_center_id') else None
        b.recurrence = request.form.get('recurrence', 'nenhuma')
        b.notes = request.form.get('notes', '').strip()
        if b.status in ('atrasado', 'pendente'):
            b.status = 'pendente' if b.due_date >= date.today() else 'atrasado'
        db.session.commit()
        flash('Lembrete atualizado!', 'success')
        return redirect(url_for('lembretes'))
    return render_template('lembretes/form.html', lembrete=b,
                           categorias=categorias, centros=centros, hoje=date.today())

@app.route('/lembretes/<int:id>/pagar', methods=['POST'])
@login_required
def lembrete_pagar(id):
    b = BillReminder.query.get_or_404(id)
    if not current_user.can_edit():
        flash('Sem permissão.', 'danger')
        return redirect(url_for('lembretes'))
    data_pgto_str = request.form.get('data_pagamento', '')
    data_pgto = datetime.strptime(data_pgto_str, '%Y-%m-%d').date() if data_pgto_str else date.today()
    # Cria lançamento automaticamente
    t = Transaction(
        date=data_pgto,
        description=f'[Lembrete] {b.description}',
        category_id=b.category_id,
        cost_center_id=b.cost_center_id,
        value=b.value,
        type='saida',
        status='realizado',
        notes=b.notes,
        user_id=current_user.id
    )
    db.session.add(t)
    db.session.flush()
    b.status = 'pago'
    b.paid_at = datetime.utcnow()
    b.transaction_id = t.id
    # Se recorrente, cria próximo lembrete
    if b.recurrence == 'mensal':
        proximo = BillReminder(
            description=b.description, value=b.value,
            due_date=b.due_date + relativedelta(months=1),
            category_id=b.category_id, cost_center_id=b.cost_center_id,
            recurrence=b.recurrence, status='pendente',
            notes=b.notes, user_id=current_user.id
        )
        db.session.add(proximo)
    elif b.recurrence == 'semanal':
        proximo = BillReminder(
            description=b.description, value=b.value,
            due_date=b.due_date + timedelta(weeks=1),
            category_id=b.category_id, cost_center_id=b.cost_center_id,
            recurrence=b.recurrence, status='pendente',
            notes=b.notes, user_id=current_user.id
        )
        db.session.add(proximo)
    db.session.commit()
    flash('Conta marcada como paga! Lançamento registrado automaticamente.', 'success')
    return redirect(url_for('lembretes'))

@app.route('/lembretes/<int:id>/cancelar', methods=['POST'])
@login_required
def lembrete_cancelar(id):
    b = BillReminder.query.get_or_404(id)
    b.status = 'cancelado'
    db.session.commit()
    flash('Lembrete cancelado.', 'success')
    return redirect(url_for('lembretes'))

@app.route('/lembretes/<int:id>/excluir', methods=['POST'])
@login_required
def lembrete_excluir(id):
    b = BillReminder.query.get_or_404(id)
    db.session.delete(b)
    db.session.commit()
    flash('Lembrete excluído.', 'success')
    return redirect(url_for('lembretes'))

# ─── USUÁRIOS ────────────────────────────────────────────────────────────────

@app.route('/usuarios')
@login_required
@admin_required
def usuarios():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('usuarios/index.html', users=users)

@app.route('/usuarios/novo', methods=['GET', 'POST'])
@login_required
@admin_required
def usuario_novo():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'viewer')
        if User.query.filter_by(email=email).first():
            flash('Este e-mail já está cadastrado.', 'danger')
        else:
            u = User(name=name, email=email,
                     password_hash=generate_password_hash(password, method='pbkdf2:sha256'),
                     role=role, active=True)
            db.session.add(u)
            db.session.commit()
            flash(f'Usuário {name} criado com sucesso!', 'success')
            return redirect(url_for('usuarios'))
    return render_template('usuarios/form.html', usuario=None)

@app.route('/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def usuario_editar(id):
    u = User.query.get_or_404(id)
    if request.method == 'POST':
        u.name = request.form.get('name', '').strip()
        u.email = request.form.get('email', '').strip()
        u.role = request.form.get('role', 'viewer')
        nova_senha = request.form.get('password', '')
        if nova_senha:
            u.password_hash = generate_password_hash(nova_senha, method='pbkdf2:sha256')
        db.session.commit()
        flash('Usuário atualizado com sucesso!', 'success')
        return redirect(url_for('usuarios'))
    return render_template('usuarios/form.html', usuario=u)

@app.route('/usuarios/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def usuario_toggle(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        flash('Você não pode desativar sua própria conta.', 'warning')
    else:
        u.active = not u.active
        db.session.commit()
        flash(f'Usuário {"ativado" if u.active else "desativado"}.', 'success')
    return redirect(url_for('usuarios'))

@app.route('/usuarios/<int:id>/excluir', methods=['POST'])
@login_required
@admin_required
def usuario_excluir(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'warning')
    elif u.transactions:
        flash('Usuário possui lançamentos vinculados. Desative-o em vez de excluir.', 'warning')
    else:
        db.session.delete(u)
        db.session.commit()
        flash('Usuário excluído.', 'success')
    return redirect(url_for('usuarios'))

# ─── PERFIL ───────────────────────────────────────────────────────────────────

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    if request.method == 'POST':
        acao = request.form.get('acao')
        if acao == 'dados':
            current_user.name = request.form.get('name', '').strip()
            novo_email = request.form.get('email', '').strip()
            if novo_email != current_user.email:
                if User.query.filter_by(email=novo_email).first():
                    flash('Este e-mail já está em uso.', 'danger')
                    return redirect(url_for('perfil'))
                current_user.email = novo_email
            db.session.commit()
            flash('Dados atualizados com sucesso!', 'success')
        elif acao == 'senha':
            senha_atual = request.form.get('senha_atual', '')
            nova_senha = request.form.get('nova_senha', '')
            confirmar = request.form.get('confirmar_senha', '')
            if not check_password_hash(current_user.password_hash, senha_atual):
                flash('Senha atual incorreta.', 'danger')
            elif nova_senha != confirmar:
                flash('As senhas não coincidem.', 'danger')
            elif len(nova_senha) < 6:
                flash('A nova senha deve ter pelo menos 6 caracteres.', 'danger')
            else:
                current_user.password_hash = generate_password_hash(nova_senha, method='pbkdf2:sha256')
                db.session.commit()
                flash('Senha alterada com sucesso!', 'success')
        return redirect(url_for('perfil'))
    return render_template('perfil.html')

# ─── REDEFINIÇÃO DE SENHA ─────────────────────────────────────────────────────

@app.route('/esqueci-senha', methods=['GET', 'POST'])
def esqueci_senha():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    token_gerado = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        u = User.query.filter_by(email=email, active=True).first()
        if u:
            # Invalidar tokens anteriores
            PasswordResetToken.query.filter_by(user_id=u.id, used=False).update({'used': True})
            token = PasswordResetToken(user_id=u.id)
            db.session.add(token)
            db.session.commit()
            token_gerado = token.token
        else:
            flash('E-mail não encontrado.', 'danger')
    return render_template('esqueci_senha.html', token_gerado=token_gerado)

@app.route('/redefinir-senha/<token>', methods=['GET', 'POST'])
def redefinir_senha(token):
    t = PasswordResetToken.query.filter_by(token=token, used=False).first()
    if not t:
        flash('Link inválido ou já utilizado.', 'danger')
        return redirect(url_for('login'))
    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')
        if nova_senha != confirmar:
            flash('As senhas não coincidem.', 'danger')
        elif len(nova_senha) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'danger')
        else:
            t.user.password_hash = generate_password_hash(nova_senha, method='pbkdf2:sha256')
            t.used = True
            db.session.commit()
            flash('Senha redefinida com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
    return render_template('redefinir_senha.html', token=token, user=t.user)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, host='0.0.0.0', port=port)
